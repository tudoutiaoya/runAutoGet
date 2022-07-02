from openpyxl import load_workbook
import time
import requests
import json

list = {} #学生列表

# 获取 学号
wb = load_workbook('跑步次数.xlsx')

sheet = wb.active
prj = sheet.columns
prjTuple = tuple(prj)
firstCol = prjTuple[0] #第一列
endCol = prjTuple[-1] #最后一列

numStu = len(firstCol)-1

for i in range(numStu):
    list[firstCol[i+1].value] = ''

wb.close()

#爬虫
cookie = input("请输入cookie")
# 输入时间
beginAt = input("请输入开始时间%Y-%m-%d")
stopAt = input("请输入结束时间%Y-%m-%d")

beginTimeArray = time.strptime(beginAt + ' 0:0:0', "%Y-%m-%d %H:%M:%S")
beginTimeStamp = str(int(time.mktime(beginTimeArray))) + '000'

stopTimeArray = time.strptime(stopAt + ' 23:59:59', "%Y-%m-%d %H:%M:%S")
stopTimeStamp = str(int(time.mktime(stopTimeArray))) + '000'


#输入表格日期抽出
beginStamp = float(int(beginTimeStamp)/1000)
beginArray = time.localtime(beginStamp)
beginDate = time.strftime("%m/%d",beginArray)

stopStamp = float(int(stopTimeStamp)/1000)
stopArray = time.localtime(stopStamp)
stopDate = time.strftime("%m/%d",stopArray)
# print(beginTimeStamp)
# print(stopTimeStamp)

url = "https://runadmin.iydsj.com/background/campusadminapi/v1/runnningmanager/list"
headers = {
        "Accept": "application/json, text/plain, */*",
        "Accept-Encoding": "gzip, deflate, br",
        "Accept-Language": "zh-CN,zh;q=0.9",
        "Connection": "keep-alive",
        "Content-Length": "146",
        "Content-Type": "application/json;charset=UTF-8",
        "Cookie": cookie,
        "Host": "runadmin.iydsj.com",
        "Origin": "https://runadmin.iydsj.com",
        "Referer": "https://runadmin.iydsj.com/",
        "rootUnid": "8000",
        "sec-ch-ua": 'Not A;Brand";v="99", "Chromium";v="99", "Google Chrome";v="99"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": "Windows",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-origin",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.74 Safari/537.36",
        "X-Requested-With": "XMLHttpRequest"
    }
i=0
for stu in list:
    data = {
        "beginAt": beginTimeStamp,
        "cid": "0",
        "keyWords": stu,
        "mid": "22120",
        "pageNum": "1",
        "pageSize": "100",
        "sid": "4301",
        "stopAt": stopTimeStamp,
        "verifyStatus": "0"
    }
    json.encoder.FLOAT_REPR = lambda x: format(x, '.3f')
    response = requests.post(url=url,headers=headers,json=data)
    content = response.text
    obj = json.loads(content)
    # print(obj)
    # print(isinstance(obj, dict))
    # print(obj.get("data").get("recordList"))
    recordList = obj.get("data").get("recordList")
    # print(len(recordList))
    if stu == "2020011179":
        print(obj)
        print(recordList)


    # 查找指定日期的次数
    success = 0  # 正常
    # failure = 0  # 异常
    dayDate = {}    #当天的数据
    tmp = ''
    for record in recordList:
        status = record.get("exceptionStatus")      #状态
        length = record.get("length")               #里程
        appealStatus = record.get("appealStatus")   #审核状态
        speed = record.get("speed")  #配速
        avgStepFreq = record.get("avgStepFreq")  #步频
        duration = int(record.get("duration"))
        ddlTime = int(int(record.get("beginAt")) / 1000) + duration
        timeArray = time.localtime(ddlTime)
        ddlTime_fen_shi_miao = time.strftime("%H:%M:%S", timeArray)
        # print(ddlTime_fen_shi_miao)

        #获取日期
        beginAt = record.get("beginAt")
        timeStamp = float(beginAt / 1000)  #毫秒时间戳转为秒级时间戳
        timeArray = time.localtime(timeStamp)  # float变为时间戳
        ntCtime_str = time.strftime("%Y-%m-%d", timeArray)  # 时间戳转成Y-M-D的str

        if (tmp != ntCtime_str):
            dayDate[ntCtime_str] = 0
            tmp = ntCtime_str


        # print(status)0:不通过(代表正常)  1:通过(代表异常)
        # 正常 并且 大于3公里才记录
        if status == 0 and length >= 400 and appealStatus == 0 and (speed>= 4 and speed<=10) and (avgStepFreq >= 60 and avgStepFreq <=240) and ("05:00:00" <= ddlTime_fen_shi_miao <= "23:00:00"):
            # success += 1
            dayDate[ntCtime_str] += length
            # print(length)
    # print("success: " + str(success))
    # print("failure: " + str(failure))
    # list[stu] = success
    success = 0
    for date in dayDate:
        if dayDate[date] >= 3000:
            success += 1
    # if stu == "2020011179":
    #     print(success)

    list[stu] = success
    if i%15 == 0:
        print("爬取进为: {:.2%}".format(i/len(list)))
    i +=1

# print(list)
#写入文件

wb = load_workbook('跑步次数.xlsx')

sheet = wb.active

columns = sheet.max_column

# print(columns)
str = beginDate + '--' + stopDate
sheet.cell(1,columns+1,str)
i = 0
for x in list:
    sheet.cell(i+2,columns+1,list[x])
    if i%15 == 0:
        print("保存进为: {:.2%}".format(i/len(list)))
    i += 1

wb.save("跑步次数.xlsx")










